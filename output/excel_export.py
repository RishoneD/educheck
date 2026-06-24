import io
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


_RED    = PatternFill('solid', fgColor='FFCCCC')
_YELLOW = PatternFill('solid', fgColor='FFFACC')
_GREEN  = PatternFill('solid', fgColor='CCFFCC')
_HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')
_BOLD   = Font(bold=True)
_RTL    = Alignment(horizontal='right', vertical='center', wrap_text=True)


COLS_BY_TEACHER = ['תלמיד', 'מקצוע', 'חומרה', 'תיאור הבעיה', 'פרטים']
COLS_BY_STUDENT = ['תלמיד', 'מקצוע', 'מורה', 'חומרה', 'תיאור הבעיה', 'פרטים']


def _write_sheet(ws, headers: list[str], rows: list[list], title: str) -> None:
    ws.title = title[:31]  # Excel מגביל שמות גיליונות ל-31 תווים
    ws.sheet_view.rightToLeft = True

    # כותרות
    ws.append(headers)
    for cell in ws[1]:
        cell.fill   = _HEADER_FILL
        cell.font   = _BOLD
        cell.alignment = _RTL

    for row_data in rows:
        ws.append(row_data)
        row_idx = ws.max_row
        sev = row_data[headers.index('חומרה')]
        fill = _RED if sev == '❌' else (_YELLOW if sev == '⚠️' else _GREEN)
        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.alignment = _RTL

    # רוחב עמודות אוטומטי
    for i, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = max(
            (len(str(ws.cell(row=r, column=i).value or '')) for r in range(1, ws.max_row + 1)),
            default=len(col_name),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def build_excel_report(findings: list[dict], class_name: str) -> bytes:
    """בונה קובץ Excel בזיכרון עם שני גיליונות ראשיים + גיליון לכל מורה."""
    wb = Workbook()
    wb.remove(wb.active)  # מסיר גיליון ברירת מחדל

    # ========== גיליון "לפי מורה" ==========
    by_teacher: dict[str, list] = defaultdict(list)
    for f in findings:
        by_teacher[f['teacher']].append(f)

    # גיליון סיכום לפי מורה
    ws_teacher = wb.create_sheet('לפי מורה')
    rows_t = [
        [f['student'], f['subject'], f['severity'], f['message'], f['details']]
        for teacher_findings in sorted(by_teacher.items(), key=lambda x: x[0])
        for f in sorted(teacher_findings[1], key=lambda x: (x['subject'], x['student']))
    ]
    _write_sheet(ws_teacher, COLS_BY_TEACHER, rows_t, 'לפי מורה')

    # גיליון נפרד לכל מורה
    for teacher, items in sorted(by_teacher.items()):
        safe_name = (teacher or 'לא ידוע')[:31]
        ws = wb.create_sheet(safe_name)
        rows = [
            [f['student'], f['subject'], f['severity'], f['message'], f['details']]
            for f in sorted(items, key=lambda x: (x['subject'], x['student']))
        ]
        _write_sheet(ws, COLS_BY_TEACHER, rows, safe_name)

    # ========== גיליון "לפי תלמיד" ==========
    ws_student = wb.create_sheet('לפי תלמיד')
    rows_s = [
        [f['student'], f['subject'], f['teacher'], f['severity'], f['message'], f['details']]
        for f in sorted(findings, key=lambda x: (x['student'], x['subject']))
    ]
    _write_sheet(ws_student, COLS_BY_STUDENT, rows_s, 'לפי תלמיד')

    # שמירה לזיכרון
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
